import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelHeaderExtractor:
    def __init__(self, input_file, output_file_path=None):
        """
        初始化Excel表头提取器

        Args:
            input_file (str): 输入Excel文件路径
            output_file (str): 输出Excel文件路径，如果为None则自动生成
        """
        self.input_file = input_file
        self.output_path = output_file_path
        self.file_name = ''


    def detect_header_rows(self, sheet, max_check_rows=10):
        """
        自动检测表头行数

        Args:
            sheet: openpyxl worksheet对象
            max_check_rows (int): 最大检查行数

        Returns:
            int: 表头行数
        """
        header_rows = 1

        # 检查合并单元格来判断表头行数
        merged_ranges = sheet.merged_cells.ranges
        max_row_in_merged = 0

        for merged_range in merged_ranges:
            if merged_range.min_row <= max_check_rows:
                max_row_in_merged = max(max_row_in_merged, merged_range.max_row)

        if max_row_in_merged > 0:
            header_rows = max_row_in_merged

        # 检查前几行的数据类型变化
        for row in range(2, min(max_check_rows + 1, sheet.max_row + 1)):
            current_row_has_text = False
            next_row_has_numbers = False

            # 检查当前行是否主要是文本
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    current_row_has_text = True
                    break

            # 检查下一行是否开始有数字
            if row < sheet.max_row:
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell_value = sheet.cell(row=row + 1, column=col).value
                    if cell_value and isinstance(cell_value, (int, float)):
                        next_row_has_numbers = True
                        break

            # 如果当前行是文本，下一行开始有数字，可能是表头结束
            if current_row_has_text and next_row_has_numbers:
                header_rows = row
                break

        return header_rows

    def copy_cell_style(self, source_cell, target_cell):
        """复制单元格样式"""
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )

        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )

        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )

        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )

    def extract_headers(self, sheet_name=None, header_rows=None):
        """
        提取Excel表头

        Args:
            sheet_name (str): 工作表名称，如果为None则使用第一个工作表
            header_rows (int): 表头行数，如果为None则自动检测

        Returns:
            bool: 是否成功提取
        """
        try:
            # 打开原始Excel文件
            source_wb = openpyxl.load_workbook(self.input_file)

            # 选择工作表
            if sheet_name:
                if sheet_name not in source_wb.sheetnames:
                    print(f"工作表 '{sheet_name}' 不存在。可用工作表: {source_wb.sheetnames}")
                    return False
                source_sheet = source_wb[sheet_name]
            else:
                source_sheet = source_wb.active
                sheet_name = source_sheet.title

            print(f"正在处理工作表: {sheet_name}")

            # 检测表头行数
            if header_rows is None:
                header_rows = self.detect_header_rows(source_sheet)

            print(f"检测到表头行数: {header_rows}")

            # 创建新的工作簿
            target_wb = openpyxl.Workbook()
            target_sheet = target_wb.active
            target_sheet.title = sheet_name

            # 复制表头数据和样式
            for row in range(1, header_rows + 1):
                for col in range(1, source_sheet.max_column + 1):
                    source_cell = source_sheet.cell(row=row, column=col)
                    target_cell = target_sheet.cell(row=row, column=col)

                    # 复制值
                    target_cell.value = source_cell.value

                    # 复制样式
                    self.copy_cell_style(source_cell, target_cell)

            # 复制合并单元格
            for merged_range in source_sheet.merged_cells.ranges:
                if merged_range.min_row <= header_rows:
                    # 确保合并范围不超出表头行
                    max_row = min(merged_range.max_row, header_rows)
                    new_range = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:{get_column_letter(merged_range.max_col)}{max_row}"
                    target_sheet.merge_cells(new_range)

            # 复制列宽
            for col in range(1, source_sheet.max_column + 1):
                col_letter = get_column_letter(col)
                if col_letter in source_sheet.column_dimensions:
                    target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

            # 复制行高
            for row in range(1, header_rows + 1):
                if row in source_sheet.row_dimensions:
                    target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

            # 保存新文件
            if not os.path.exists(self.output_path):
                os.makedirs(self.output_path)
            target_wb.save(str(Path(self.output_path, self.file_name)))

            # 关闭工作簿
            source_wb.close()
            target_wb.close()

            return True

        except Exception as e:
            print(f"提取表头时发生错误: {str(e)}")
            return False

    def extract_all_sheets(self, header_rows=None):
        """
        提取所有工作表的表头

        Args:
            header_rows (int): 表头行数，如果为None则自动检测

        Returns:
            bool: 是否成功提取
        """
        try:
            # 打开原始Excel文件
            source_wb = openpyxl.load_workbook(self.input_file)

            # 创建新的工作簿
            target_wb = openpyxl.Workbook()
            target_wb.remove(target_wb.active)  # 删除默认工作表

            for sheet_name in source_wb.sheetnames:
                print(f"正在处理工作表: {sheet_name}")

                source_sheet = source_wb[sheet_name]

                # 检测表头行数
                current_header_rows = header_rows
                if current_header_rows is None:
                    current_header_rows = self.detect_header_rows(source_sheet)

                print(f"工作表 '{sheet_name}' 检测到表头行数: {current_header_rows}")

                # 创建新工作表
                target_sheet = target_wb.create_sheet(title=sheet_name)

                # 复制表头数据和样式
                for row in range(1, current_header_rows + 1):
                    for col in range(1, source_sheet.max_column + 1):
                        source_cell = source_sheet.cell(row=row, column=col)
                        target_cell = target_sheet.cell(row=row, column=col)

                        # 复制值
                        target_cell.value = source_cell.value

                        # 复制样式
                        self.copy_cell_style(source_cell, target_cell)

                # 复制合并单元格
                for merged_range in source_sheet.merged_cells.ranges:
                    if merged_range.min_row <= current_header_rows:
                        max_row = min(merged_range.max_row, current_header_rows)
                        new_range = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:{get_column_letter(merged_range.max_col)}{max_row}"
                        target_sheet.merge_cells(new_range)

                # 复制列宽
                for col in range(1, source_sheet.max_column + 1):
                    col_letter = get_column_letter(col)
                    if col_letter in source_sheet.column_dimensions:
                        target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[
                            col_letter].width

                # 复制行高
                for row in range(1, current_header_rows + 1):
                    if row in source_sheet.row_dimensions:
                        target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

            # 保存新文件
            target_wb.save(str(Path(self.output_path,self.file_name)))

            # 关闭工作簿
            source_wb.close()
            target_wb.close()

            return True

        except Exception as e:
            print(f"提取表头时发生错误: {str(e)}")
            return False


def example():
    """主函数 - 使用示例"""
    # 使用示例
    input_file = "your_input_file.xlsx"  # 请替换为你的输入文件路径

    # 创建提取器实例
    extractor = ExcelHeaderExtractor(input_file)

    # 方法1: 提取指定工作表的表头（自动检测表头行数）
    success = extractor.extract_headers()

    # 方法2: 提取指定工作表的表头（手动指定表头行数）
    # success = extractor.extract_headers(sheet_name="Sheet1", header_rows=3)

    # 方法3: 提取所有工作表的表头
    # success = extractor.extract_all_sheets()

    if success:
        print("表头提取完成！")
    else:
        print("表头提取失败！")