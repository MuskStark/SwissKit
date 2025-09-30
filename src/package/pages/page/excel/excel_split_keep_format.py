import shutil
from enum import Enum
from pathlib import Path
from typing import Any, Dict, Optional, Callable

import pandas as pd

from ....enums.progress_status_enums import (ProgressStatus)


class ExcelSplitConfig:
    def __init__(self):
        self.split_model: int = 0
        self.split_sub_model: str = ""
        self.split_config: list[dict[str, Any]] = []


class ExcelSplitKM:

    def __init__(self, progress_callback: Optional[Callable[[Enum, str], None]] = None):
        """
        初始化Excel拆分处理器

        Args:
            progress_callback: 进度回调函数，接收(status, message)参数
        """
        self.progress_callback = progress_callback

    def _update_progress(self, status: Enum, message: str):
        """更新进度"""
        if self.progress_callback:
            self.progress_callback(status, message)

    def split_keep_format(self, excel, excel_split_config: ExcelSplitConfig,
                          output_folder_path: str) -> bool:
        """
        根据配置拆分Excel文件并保持格式

        Args:
            excel: Excel对象
            excel_split_config: 拆分配置
            output_folder_path: 输出文件夹路径

        Returns:
            bool: 拆分是否成功
        """
        # model-0: 简单拆分  model-1: 多表头拆分
        if excel_split_config.split_model == 0:
            if excel_split_config.split_sub_model == 'sheet':
                return self._split_file_by_sheet(excel, output_folder_path)
            elif excel_split_config.split_sub_model == 'col':
                # 需要传入拆分列信息和sheet名称
                if excel_split_config.split_config:
                    sheet_name = excel_split_config.split_config[0].get('sheet_name', None)
                    split_column = excel_split_config.split_config[0].get('split_column', '')
                    return self._split_file_by_col(excel, output_folder_path, split_column, sheet_name)
            elif excel_split_config.split_sub_model == 'multi_col':
                # 多Sheet按列拆分
                if excel_split_config.split_config:
                    config = excel_split_config.split_config[0]
                    return self._split_multi_sheets_by_col(
                        excel,
                        output_folder_path,
                        config['selected_sheets'],
                        config['split_column'],
                        config['result_dict']
                    )
        elif excel_split_config.split_model == 1:
            if excel_split_config.split_sub_model == 'multiple':
                return self._split_multiple_headers_excel(excel, output_folder_path, excel_split_config.split_config)

        return False

    def _split_file_by_sheet(self, excel, output_folder_path: str) -> bool:
        """按工作表拆分Excel文件"""
        try:
            from openpyxl import load_workbook, Workbook

            self._update_progress(ProgressStatus.LOADING, '开始按Sheet拆分并保持格式')
            original_wb = load_workbook(excel.file_path)
            file_name = Path(excel.file_path).stem

            for sheet_name in excel.sheets.keys():
                self._update_progress(ProgressStatus.LOADING, f'生成格式化文件: {sheet_name}')

                # Create new workbook with single sheet
                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.title = sheet_name

                # Copy the original sheet content and formatting
                original_ws = original_wb[sheet_name]
                self._copy_worksheet_complete(original_ws, new_ws)

                # Save the new file
                output_file = Path(output_folder_path, f"{file_name}_{sheet_name}.xlsx")
                new_wb.save(output_file)
                new_wb.close()

            original_wb.close()
            self._update_progress(ProgressStatus.SUCCESS, '按Sheet格式保持拆分完成')
            return True

        except Exception as e:
            self._update_progress(ProgressStatus.ERROR, f'按Sheet格式保持拆分失败: {str(e)}')
            return False

    def _split_file_by_col(self, excel,
                           output_folder_path: str, split_column: str, sheet_name: str = None) -> bool:
        """按指定列的值拆分Excel文件"""
        try:
            from openpyxl import load_workbook, Workbook

            self._update_progress(ProgressStatus.LOADING, '开始按列拆分并保持格式')
            original_wb = load_workbook(excel.file_path)
            file_name = Path(excel.file_path).stem

            # 如果指定了sheet_name，只处理该sheet，否则处理所有sheet
            sheets_to_process = [sheet_name] if sheet_name else list(excel.sheets.keys())

            for current_sheet_name in sheets_to_process:
                original_ws = original_wb[current_sheet_name]
                sheet_obj = excel.sheets[current_sheet_name]

                # 找到分割列的索引
                split_col_idx = None
                for idx, col in enumerate(sheet_obj.columns):
                    if col == split_column:
                        split_col_idx = idx + 1  # openpyxl uses 1-based indexing
                        break

                if split_col_idx is None:
                    continue

                # 获取该列的所有唯一值
                unique_values = set()
                for row in original_ws.iter_rows(min_row=2, min_col=split_col_idx, max_col=split_col_idx):
                    if row[0].value:
                        unique_values.add(str(row[0].value))

                # 为每个唯一值创建新文件
                for value in unique_values:
                    self._update_progress(ProgressStatus.LOADING, f'生成文件: {current_sheet_name}_{value}')

                    new_wb = Workbook()
                    new_ws = new_wb.active
                    new_ws.title = current_sheet_name

                    # 复制表头
                    self._copy_header_row(original_ws, new_ws)

                    # 复制符合条件的数据行
                    self._copy_filtered_rows(original_ws, new_ws, split_col_idx, value)

                    # 保存文件
                    safe_value = "".join(c for c in value if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    if sheet_name:
                        # 单sheet拆分，不需要在文件名中加sheet名
                        output_file = Path(output_folder_path, f"{file_name}_{safe_value}.xlsx")
                    else:
                        # 多sheet拆分，需要在文件名中加sheet名
                        output_file = Path(output_folder_path, f"{file_name}_{current_sheet_name}_{safe_value}.xlsx")
                    new_wb.save(output_file)
                    new_wb.close()

            original_wb.close()
            self._update_progress(ProgressStatus.SUCCESS, '按列格式保持拆分完成')
            return True

        except Exception as e:
            self._update_progress(ProgressStatus.ERROR, f'按列格式保持拆分失败: {str(e)}')
            return False

    def _split_multi_sheets_by_col(self, excel,
                                   output_folder_path: str,
                                   selected_sheets: list,
                                   split_column: str,
                                   result_dict: dict) -> bool:
        """多Sheet按列拆分，保持格式"""
        try:
            from openpyxl import load_workbook

            self._update_progress(ProgressStatus.LOADING, '开始多Sheet按列拆分并保持格式')
            file_stem = Path(excel.file_path).stem

            for group_name, sheets_data in result_dict.items():
                self._update_progress(ProgressStatus.LOADING, f'生成文件: {file_stem}_{group_name}.xlsx')

                wb = load_workbook(excel.file_path)

                # 删除未选中的Sheet
                for ws_name in list(wb.sheetnames):
                    if ws_name not in selected_sheets:
                        del wb[ws_name]

                # 处理每个Sheet的数据
                for sheet_name, group_data in sheets_data.items():
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]

                        # 找到拆分列的索引
                        col_idx = None
                        for col in range(1, ws.max_column + 1):
                            if ws.cell(1, col).value == split_column:
                                col_idx = col
                                break

                        if col_idx:
                            # 删除不属于该组的行（从后往前删除）
                            rows_to_keep = set(group_data.index + 2)  # +2 因为有表头行，且DataFrame索引从0开始
                            for row in range(ws.max_row, 1, -1):
                                if row not in rows_to_keep:
                                    ws.delete_rows(row)

                output_path = Path(output_folder_path, f"{file_stem}_{group_name}.xlsx")
                wb.save(output_path)
                wb.close()

            self._update_progress(ProgressStatus.SUCCESS, '多Sheet按列格式保持拆分完成')
            return True

        except Exception as e:
            self._update_progress(ProgressStatus.ERROR, f'多Sheet按列格式保持拆分失败: {str(e)}')
            return False

    def _split_multiple_headers_excel(self, excel,
                                      output_folder_path: str,
                                      split_config: list[dict[str, Any]]) -> bool:
        """
        根据多表头配置拆分Excel文件

        Args:
            excel: Excel对象
            output_folder_path: 输出文件夹路径
            split_config: 拆分配置列表，每个配置包含:
                - sheet_name: 工作表名称
                - header_rows: 表头行数
                - split_column_index: 拆分列索引(0-based)

        Returns:
            bool: 拆分是否成功
        """
        temp_folder = None
        try:
            # 生成模板文件配置
            split_config_dic = self._generate_template_config(excel, output_folder_path, split_config)
            if not split_config_dic:
                return False

            temp_folder = Path(output_folder_path, 'tmp')

            self._update_progress(ProgressStatus.LOADING, '开始分析数据')

            # 读取和分组数据
            df_group_dic = {}
            for sheet_name, config in split_config_dic.items():
                sheet_object = excel.sheets.get(sheet_name)

                # 读取数据（跳过表头行）
                df_data = pd.read_excel(
                    excel.file_path,
                    header=None,
                    sheet_name=sheet_name,
                    skiprows=config['header_rows']
                )

                # 按指定列分组
                df_group_dic[sheet_name] = df_data.groupby(
                    df_data.columns[config['split_column_index']]
                )

            # 重新组织数据结构
            result_dic = {}
            for sheet, df_group in df_group_dic.items():
                for group_key, data in df_group:
                    if group_key not in result_dic:
                        result_dic[group_key] = {sheet: data}
                    else:
                        result_dic[group_key][sheet] = data

            self._update_progress(ProgressStatus.LOADING, '开始生成文件')

            # 生成最终文件
            total_files = len(result_dic)
            current_file = 0

            for group_key, sheets_data in result_dic.items():
                current_file += 1

                # 处理文件名中的特殊字符
                safe_key = str(group_key).replace('/', '-').replace('\\', '-').replace(':', '-')
                file_name = f"{Path(excel.file_path).stem}_{safe_key}.xlsx"

                self._update_progress(ProgressStatus.LOADING,
                                      f'生成文件 ({current_file}/{total_files}): {file_name}')

                output_file_path = Path(output_folder_path, file_name)

                # 处理多个sheet的情况
                if len(sheets_data) == 1:
                    # 只有一个sheet，直接使用模板
                    sheet_name = list(sheets_data.keys())[0]
                    template_path = split_config_dic[sheet_name]['template_file_path']
                    data_df = sheets_data[sheet_name]
                    header_rows = split_config_dic[sheet_name]['header_rows']

                    self._write_data_with_format(template_path, data_df, sheet_name, header_rows, output_file_path)

                else:
                    # 多个sheet，需要合并到一个文件
                    self._create_multi_sheet_file(sheets_data, split_config_dic, output_file_path)

            self._update_progress(ProgressStatus.SUCCESS, f'拆分完成，共生成 {total_files} 个文件')
            return True

        except Exception as e:
            self._update_progress(ProgressStatus.ERROR, f'拆分失败: {str(e)}')
            return False

        finally:
            # 清理临时文件
            if temp_folder and temp_folder.exists():
                try:
                    shutil.rmtree(temp_folder)
                except Exception as e:
                    print(f"清理临时文件失败: {e}")

    def _generate_template_config(self, excel,
                                  output_folder_path: str,
                                  split_config: list[dict[str, Any]]) -> Optional[Dict[str, Any]]:
        """
        生成基于用户分割配置的模板文件配置

        Args:
            excel: Excel对象
            output_folder_path: 输出文件夹路径
            split_config: 拆分配置列表

        Returns:
            dict: 包含每个选中sheet配置信息的字典，包括模板文件路径
        """
        self._update_progress(ProgressStatus.LOADING, '开始拆分前准备')

        split_config_dic = {}

        # 解析用户配置
        for config in split_config:
            try:
                sheet_name = config['sheet_name']
                header_rows = int(config['header_rows'])
                split_column_index = int(config['split_column_index'])

                split_config_dic[sheet_name] = {
                    "header_rows": header_rows,
                    "split_column_index": split_column_index
                }
            except (ValueError, KeyError) as e:
                self._update_progress(ProgressStatus.ERROR, f"配置错误: {str(e)}")
                return None

        if not split_config_dic:
            self._update_progress(ProgressStatus.ERROR, "请至少选择一个Sheet进行拆分")
            return None

        # 创建临时文件夹
        tmp_folder = Path(output_folder_path, 'tmp')
        tmp_folder.mkdir(parents=True, exist_ok=True)

        # 为每个选中的sheet创建带完整格式的模板文件
        from openpyxl import load_workbook, Workbook

        try:
            source_wb = load_workbook(excel.file_path)

            for selected_sheet_name in split_config_dic.keys():
                self._update_progress(ProgressStatus.LOADING, f'准备模板: {selected_sheet_name}')

                # 创建新的工作簿作为模板
                template_wb = Workbook()
                template_ws = template_wb.active
                template_ws.title = selected_sheet_name

                source_ws = source_wb[selected_sheet_name]

                # 复制表头部分（包含完整格式）
                header_rows = split_config_dic[selected_sheet_name]['header_rows']
                self._copy_header_with_format(source_ws, template_ws, header_rows)

                # 保存模板文件
                template_file_name = f"{Path(excel.file_path).stem}_{selected_sheet_name}_template.xlsx"
                template_file_path = tmp_folder / template_file_name
                template_wb.save(template_file_path)
                template_wb.close()

                split_config_dic[selected_sheet_name]['template_file_path'] = template_file_path

            source_wb.close()
            self._update_progress(ProgressStatus.LOADING, '模板准备完成')
            return split_config_dic

        except Exception as e:
            self._update_progress(ProgressStatus.ERROR, f'模板生成失败: {str(e)}')
            return None

    def _create_multi_sheet_file(self, sheets_data: Dict[str, pd.DataFrame],
                                 split_config_dic: Dict[str, Any],
                                 output_file_path: Path):
        """创建包含多个工作表的文件"""
        from openpyxl import Workbook, load_workbook

        final_wb = Workbook()
        final_wb.remove(final_wb.active)  # 移除默认sheet

        for sheet_name, data_df in sheets_data.items():
            template_path = split_config_dic[sheet_name]['template_file_path']
            header_rows = split_config_dic[sheet_name]['header_rows']

            # 加载模板
            template_wb = load_workbook(template_path)
            template_ws = template_wb[sheet_name]

            # 在最终工作簿中创建新sheet
            final_ws = final_wb.create_sheet(sheet_name)

            # 复制模板格式
            self._copy_worksheet_complete(template_ws, final_ws)

            # 写入数据
            start_row = header_rows + 1
            for row_idx, (_, row_data) in enumerate(data_df.iterrows(), start=start_row):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = final_ws.cell(row=row_idx, column=col_idx)
                    cell.value = value

            template_wb.close()

        # 保存最终文件
        final_wb.save(output_file_path)
        final_wb.close()

    def _copy_header_row(self, source_ws, target_ws):
        """复制表头行"""
        for cell in source_ws[1]:
            target_cell = target_ws.cell(row=1, column=cell.column)
            target_cell.value = cell.value
            if cell.has_style:
                target_cell.font = cell.font.copy()
                target_cell.border = cell.border.copy()
                target_cell.fill = cell.fill.copy()
                target_cell.alignment = cell.alignment.copy()

    def _copy_filtered_rows(self, source_ws, target_ws, filter_col_idx, filter_value):
        """复制符合条件的数据行"""
        target_row = 2
        for row in source_ws.iter_rows(min_row=2):
            if str(row[filter_col_idx - 1].value) == filter_value:
                for cell in row:
                    target_cell = target_ws.cell(row=target_row, column=cell.column)
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.alignment = cell.alignment.copy()
                target_row += 1

    def _copy_worksheet_complete(self, source_ws, target_ws):
        """
        完整复制工作表包括数据、格式和结构

        Args:
            source_ws: 源openpyxl工作表
            target_ws: 目标openpyxl工作表
        """
        # 复制单元格数据和格式
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value

                # 复制所有格式
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.protection = cell.protection.copy()
                    target_cell.alignment = cell.alignment.copy()

        # 复制列宽
        for col in source_ws.column_dimensions:
            target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

        # 复制行高
        for row in source_ws.row_dimensions:
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

        # 复制合并单元格
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        # 复制工作表属性
        if hasattr(source_ws, 'sheet_properties') and hasattr(target_ws, 'sheet_properties'):
            target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor

    def _copy_header_with_format(self, source_ws, target_ws, header_rows):
        """
        复制表头行，包含完整的格式信息

        Args:
            source_ws: 源工作表
            target_ws: 目标工作表
            header_rows: 表头行数
        """
        # 复制表头行的数据和格式
        for row_idx in range(1, header_rows + 1):
            for cell in source_ws[row_idx]:
                target_cell = target_ws.cell(row=cell.row, column=cell.column)
                target_cell.value = cell.value

                # 复制完整格式
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.protection = cell.protection.copy()
                    target_cell.alignment = cell.alignment.copy()

        # 复制列宽
        for col in source_ws.column_dimensions:
            if source_ws.column_dimensions[col].width:
                target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

        # 复制表头区域的行高
        for row_idx in range(1, header_rows + 1):
            if row_idx in source_ws.row_dimensions:
                target_ws.row_dimensions[row_idx].height = source_ws.row_dimensions[row_idx].height

        # 复制表头区域的合并单元格
        for merged_range in source_ws.merged_cells.ranges:
            # 只复制在表头范围内的合并单元格
            if merged_range.min_row <= header_rows:
                target_ws.merge_cells(str(merged_range))

    def _write_data_with_format(self, template_file_path, data_df, sheet_name, header_rows, output_path):
        """
        将数据写入模板文件并保持格式

        Args:
            template_file_path: 模板文件路径
            data_df: 要写入的数据DataFrame
            sheet_name: 工作表名称
            header_rows: 表头行数
            output_path: 输出文件路径
        """
        from openpyxl import load_workbook

        # 复制模板到输出位置
        shutil.copy2(template_file_path, output_path)

        # 打开文件并写入数据
        wb = load_workbook(output_path)
        ws = wb[sheet_name]

        # 从表头行之后开始写入数据
        start_row = header_rows + 1

        # 逐行写入数据，保持原有格式
        for row_idx, (_, row_data) in enumerate(data_df.iterrows(), start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # 如果原始位置有格式，尝试继承
                if row_idx > start_row:  # 不是第一行数据
                    try:
                        # 从上一行复制格式
                        source_cell = ws.cell(row=row_idx - 1, column=col_idx)
                        if source_cell.has_style:
                            cell.font = source_cell.font.copy()
                            cell.border = source_cell.border.copy()
                            cell.fill = source_cell.fill.copy()
                            cell.number_format = source_cell.number_format
                            cell.alignment = source_cell.alignment.copy()
                    except:
                        pass

        wb.save(output_path)
        wb.close()
