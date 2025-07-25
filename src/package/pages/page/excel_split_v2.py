import os
import platform
import shutil
import subprocess
import tempfile
from pathlib import Path
from time import sleep
from typing import cast

import flet as ft
import pandas as pd
from openpyxl import load_workbook

from ..toolbox_page import ToolBoxPage
from ...components.progress_ring_components import ProgressRingComponent
from ...enums.progress_status_enums import ProgressStatus
from ...util.excel_util import ExcelHeaderExtractor


def open_folder_in_explorer(path):
    path = Path(path)
    if not path.exists():
        return False

    path_str = str(path.absolute())

    try:
        if platform.system() == 'Windows':
            os.startfile(path_str)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', path_str])
        else:  # Linux
            subprocess.run(['xdg-open', path_str])
        return True
    except Exception as e:
        return False


class ExcelSplitPageV2(ToolBoxPage):
    # 初始化
    def __init__(self, main_page: ft.Page):
        self.theme = {
            "text_color": ft.Colors.GREY_600
        }
        self.file_analyze_dic = {}
        self.page = main_page
        self.disable = True
        self.excel = None
        self.checkBox = ft.Checkbox(label='拆分后打开输出文件夹', value=True)
    # 内部类
    class ExcelObject:
        def __init__(self, file_path: str):
            self.file_path = file_path
            self.sheets = {}
        def add_sheet(self,sheet_name:str, sheet_obj):
            if sheet_name in self.sheets:
                raise ValueError(f"Sheet '{sheet_name}' already exists.")
            self.sheets[sheet_name] = sheet_obj

    class ExcelSheetObject:
        def __init__(self, sheet_name, columns, data: pd.DataFrame):
            self.sheet_name = sheet_name
            self.columns = columns
            self.df_data = data

        def __repr__(self):
            return f"ExcelSheet(sheet_name={self.sheet_name}, columns={self.columns})"


    # public
    def _load_excel_file(self, file_path_text: ft.TextField,progress: ProgressRingComponent,tab_page:ft.Tabs, advance_model:bool=False):
        """
        Load an Excel file and parse its content into the application's internal data structure.

        Summary:
        This method loads an Excel file specified by a text field, updates the UI to reflect the progress of the operation,
        and parses the Excel file's content into a structured format that can be used by the application. It supports
        both standard and advanced models for parsing the sheets within the Excel file.

        Parameters:
        file_path_text (ft.TextField): A text field containing the path to the Excel file.
        progress (ProgressRingComponent): A component to display and update the status of the loading process.
        tab_page (ft.Tabs): The tab page to be made visible after successful loading.
        advance_model (bool, optional): A flag indicating whether to use the advanced model for parsing. Defaults to False.

        Raises:
        Exception: If there is an error during the file loading or parsing process, it will be caught, and the error message
        will be displayed in the progress component.

        Returns:
        None: This function does not return any value but updates the internal state and UI components directly.
        """
        try:
            file_path = Path(file_path_text.value)
            if file_path:
                if file_path.suffix == '.xlsx':
                    excel_obj = self.ExcelObject(file_path_text.value)
                    progress.update_status(ProgressStatus.LOADING, '开始解析Excel文件')
                    if not advance_model :
                        all_sheets_dict = pd.read_excel(file_path, sheet_name=None, dtype=str)
                        for sheet_name, sheet_data in all_sheets_dict.items():
                            sheet_obj = self.ExcelSheetObject(sheet_name,sheet_data.columns.tolist(), sheet_data)
                            excel_obj.add_sheet(sheet_name, sheet_obj)
                        if self.excel is not None:
                            self.excel = None
                        self.excel = excel_obj
                    else:
                        excel_file = pd.ExcelFile(file_path)
                        for _sheet in excel_file.sheet_names:
                            sheet_obj = self.ExcelSheetObject(_sheet, None, pd.DataFrame())
                            excel_obj.add_sheet(_sheet, sheet_obj)
                    tab_page.visible=True
                    progress.update_status(ProgressStatus.SUCCESS, '完成解析')
                    self.page.update()
                else:
                    progress.update_status(ProgressStatus.ERROR,'请使用xlsx格式')
            else:
                progress.update_status(ProgressStatus.ERROR,'文件路径为空')
        except Exception as e:
            progress.update_status(ProgressStatus.ERROR, str(e))

    def _picker(self, e: ft.FilePickerResultEvent,
                file_path_text: ft.TextField,
                is_file: bool = True):
        if is_file:
            if e.files:
                if Path(e.files[0].path).suffix.lower() not in ['.xlsx']:
                    file_path_text.value = ''
                    file_path_text.error_text = '请上传xlsx文件'
                    file_path_text.update()
                else:
                    file_path_text.value = e.files[0].path
                    file_path_text.error_text = ''
                    file_path_text.update()
        else:
            if e.path:
                file_path_text.value = e.path
        self.page.update()

    def _tab_change(self, e, output_folder_path_text:ft.TextField, _tabs:ft.Tabs):
        """
        Handle tab change event to update the content of the first tab with split Excel data.

        :raises: None
        :returns: None

        :param e: The event object that triggered the tab change.
        :type e: Event
        :param output_folder_path_text: Text field containing the path to the output folder.
        :type output_folder_path_text: ft.TextField
        :param _tabs: Tabs control containing the tabs to be updated.
        :type _tabs: ft.Tabs
        """
        if e.control.selected_index == 1:
            _tabs.tabs[0].content = self._split_excel(output_folder_path_text)
            _tabs.update()
            # e.control.tabs[1].content = self._split_excel(output_folder_path_text)
            e.page.update()


    def _split_excel(self, output_folder_path_text:ft.TextField):
        """
        Splits an Excel file into multiple files based on the selected sheets and a common column.

        This method provides a user interface to select which sheets from the Excel workbook should be split,
        and it allows the user to choose a common column that will be used to group the data. After
        selecting the output folder, the selected sheets are processed and saved as separate Excel
        files, each containing data grouped by the values in the chosen column.

        :raises:
            Exception: If there is an error during the processing of the Excel file.

        :returns:
            ft.Column: A column component containing the UI elements for selecting sheets, choosing a common column, and starting the split process.
        """
        def after_select_check_box(_check_box_list_components, _drop_down:ft.Dropdown):
            _columns_set = None
            for box in check_box_list_components.controls:
                checkbox = cast(ft.Checkbox, box)
                if checkbox.value:
                    if _columns_set is None:
                        _columns_set = set(self.excel.sheets[checkbox.label].columns)
                    _columns_set = _columns_set.intersection(set(self.excel.sheets[checkbox.label].columns))
            _drop_down.options = [ft.DropdownOption(value) for value in _columns_set]
            _drop_down.update()

        def split_logic(_folder_path_text:ft.TextField, _check_box_list_components, _drop_down:ft.Dropdown, _process_ring:ProgressRingComponent):
            result_dic = {}  # 初始化为空字典而不是None
            _process_ring.update_status(ProgressStatus.LOADING, "开始拆分")

            try:
                _select_sheet = []
                for box in _check_box_list_components.controls:
                    checkbox = cast(ft.Checkbox, box)
                    if checkbox.value:
                        _select_sheet.append(checkbox.label)

                for sheet in _select_sheet:
                    df_sheet_data = self.excel.sheets[sheet].df_data
                    group = df_sheet_data.groupby(_drop_down.value)

                    for k, v in group:
                        if k not in result_dic:
                            result_dic[k] = {sheet: v}
                        else:
                            result_dic[k][sheet] = v

                for k, v in result_dic.items():
                    out_file = Path(_folder_path_text.value, Path(self.excel.file_path).stem + f'_{k}.xlsx')
                    _process_ring.update_status(ProgressStatus.LOADING, f'开始生成-{str(out_file)}')

                    with pd.ExcelWriter(out_file) as writer:
                        for _k, _v in v.items():  # 遍历字典的items()
                            _v.to_excel(writer, sheet_name=_k, index=False)

                _process_ring.update_status(ProgressStatus.SUCCESS, "完成拆分")
                open_folder_in_explorer(output_folder_path_text.value)

            except Exception as e:
                _process_ring.update_status(ProgressStatus.ERROR, str(e))

        if self.excel is None:
            return None
        else:
            check_box_list_components = ft.Row(alignment=ft.MainAxisAlignment.CENTER,expand=True)
            for sheet_name in self.excel.sheets.keys():
                check_box = ft.Checkbox(label=sheet_name, value=False)
                check_box_list_components.controls.append(check_box)
            get_columns_button = ft.ElevatedButton(text='获取可拆分列', on_click= lambda _: after_select_check_box(check_box_list_components,drop_down))
            drop_down = ft.Dropdown(options=[],width=200)
            processing_ring = ProgressRingComponent()
            start_analyze = ft.ElevatedButton(text='开始拆分', on_click=lambda _:split_logic(output_folder_path_text, check_box_list_components, drop_down, processing_ring))

            return ft.Column(
                controls=[
                    check_box_list_components,
                    ft.Row([get_columns_button],alignment=ft.MainAxisAlignment.CENTER),
                    ft.Row([drop_down],alignment=ft.MainAxisAlignment.CENTER),
                    ft.Row([start_analyze], alignment=ft.MainAxisAlignment.CENTER),
                    ft.Row([processing_ring],alignment=ft.MainAxisAlignment.CENTER)
                ],
                spacing=30,
                expand=True
            )

    def _split_multiple_headers_excel(self,output_folder_path_text:ft.TextField):
        def _generate_tmp_file(_output_folder_path_text:ft.TextField, _split_config_component:ft.Column, _processing:ProgressRingComponent):
            """
            Generates temporary files based on the split configuration provided by the user.

            Summary:
            This function prepares and generates temporary files for each selected sheet in an Excel file according to the user's specified split configuration. It validates the input, extracts headers from the sheets, and creates a temporary file for each. The process is wrapped with a progress status update mechanism.

            Args:
                _output_folder_path_text (ft.TextField): The text field containing the path to the output folder.
                _split_config_component (ft.Column): A column component holding the split configuration rows.
                _processing (ProgressRingComponent): A component to display the progress of the operation.

            Raises:
                ValueError: If the header rows or split column index are not valid integers.
                RuntimeError: If there is a failure in extracting headers, if all sheets from the source file are not parsed, or if there is an issue with the header parsing.

            Returns:
                dict: A dictionary mapping each selected sheet name to its configuration including the path to the generated temporary file.
            """
            _processing.update_status(ProgressStatus.LOADING, '开始拆分前准备')
            # Generate temporary files based on the selected sheet
            _controls = _split_config_component.controls
            split_config_dic = {}
            # Get the ft.Row component from each _split_config
            for row_control in _controls:
                row_controls = row_control.controls
                # Get config from each ft.Row component
                _check_box = row_controls[0]
                _header_rows = row_controls[1]
                _split_column_index = row_controls[2]
                # valid config
                try:
                    if _check_box.value:
                        _header_rows = int(_header_rows.value)
                        _split_column_index = int(_split_column_index.value) - 1
                        _select_sheet_name = _check_box.label
                        split_config_dic[_select_sheet_name] = {"header_index": _header_rows,
                                                                "column_index": _split_column_index}
                except ValueError as e:
                    _processing.update_status(ProgressStatus.ERROR, str(e))
                    return None


            # Generate temporary files based on the selected split configuration from the interface
            _file_path = self.excel.file_path
            extractor = ExcelHeaderExtractor(_file_path, str(Path(_output_folder_path_text.value, 'tmp')))
            for _selected_sheet_name in split_config_dic.keys():
                extractor.file_name = Path(self.excel.file_path).stem + f"_{_selected_sheet_name}_header_tmp.xlsx"
                result = extractor.extract_headers(_selected_sheet_name, split_config_dic[_selected_sheet_name]['header_index'])
                if not result:
                    raise RuntimeError('解析excel表头失败')
                split_config_dic[_selected_sheet_name]['tmp_file_path'] = Path(_output_folder_path_text.value,'tmp', extractor.file_name)
            # Verify that the number of generated temporary files matches the source file
            if split_config_dic:
                # if len(split_config_dic.keys()) != len(self.excel.sheets.keys()):
                #     raise RuntimeError('源文件Sheet页未能全部解析')
                processing_ring.update_status(ProgressStatus.LOADING, '完成拆分前准备')
                return split_config_dic
            else:
                raise RuntimeError('表头解析异常')

        def _split_logic(_output_folder_path_text:ft.TextField, _split_config_component:ft.Column, _processing:ProgressRingComponent):
            try:
                _split_config_dic = _generate_tmp_file(output_folder_path_text, _split_config_component, _processing)
                # Update ExcelObject meta
                processing_ring.update_status(ProgressStatus.LOADING,'开始拆分')
                df_group_dic = {}
                for _sheet_name in _split_config_dic.keys():
                    _sheet_object = self.excel.sheets.get(_sheet_name)
                    _sheet_object.columns = _split_config_dic[_sheet_name]['column_index']
                    _sheet_object.df_data = pd.read_excel(self.excel.file_path,header=None, sheet_name=_sheet_name, skiprows=_split_config_dic[_sheet_name]['header_index'])
                    df_group_dic[_sheet_name] = _sheet_object.df_data.groupby(_sheet_object.df_data.columns[_sheet_object.columns])
                # start split
                result_dic = {}
                for sheet, df_group in df_group_dic.items():
                    for k, data in df_group:
                        if k not in result_dic:
                            result_dic[k] = {sheet:data}
                        else:
                            result_dic[k][sheet] = data

                # Generate file
                # Copy template file to out file
                for k, v in result_dic.items():
                    if '/' in k:
                        k = k.replace('/','-')
                    file_name = Path(self.excel.file_path).stem + f'_{k}.xlsx'
                    processing_ring.update_status(ProgressStatus.LOADING, f'生成文件：{file_name}')
                    out_file_path = Path(output_folder_path_text.value, file_name)

                    temp_files = []

                    try:
                        # 为每个sheet创建临时文件
                        for _sheet in v.keys():
                            template_path = _split_config_dic.get(_sheet)['tmp_file_path']

                            # 创建临时文件
                            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
                            os.close(temp_fd)
                            temp_files.append(temp_path)

                            # 复制模板到临时文件
                            shutil.copy(template_path, temp_path)

                            # 写入数据到临时文件
                            with pd.ExcelWriter(temp_path, engine='openpyxl', mode='a',
                                                if_sheet_exists='overlay') as writer:
                                v[_sheet].to_excel(writer, sheet_name=_sheet, index=False, header=False,
                                                   startrow=_split_config_dic[_sheet]['header_index'])

                        # 合并所有临时文件到最终文件
                        final_wb = load_workbook(temp_files[0])

                        def copy_worksheet_content(source_ws, target_ws):
                            """手动复制工作表内容"""
                            # 复制单元格数据和格式
                            for row in source_ws.iter_rows():
                                for cell in row:
                                    target_cell = target_ws.cell(row=cell.row, column=cell.column)
                                    target_cell.value = cell.value

                                    # 复制格式
                                    if cell.has_style:
                                        target_cell.font = cell.font.copy()
                                        target_cell.border = cell.border.copy()
                                        target_cell.fill = cell.fill.copy()
                                        target_cell.number_format = cell.number_format
                                        target_cell.protection = cell.protection.copy()
                                        target_cell.alignment = cell.alignment.copy()

                            # 复制列宽
                            for col in source_ws.column_dimensions:
                                target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

                            # 复制行高
                            for row in source_ws.row_dimensions:
                                target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

                            # 复制合并单元格
                            for merged_range in source_ws.merged_cells.ranges:
                                target_ws.merge_cells(str(merged_range))

                        for i, temp_path in enumerate(temp_files[1:], 1):
                            temp_wb = load_workbook(temp_path)
                            sheet_name = list(v.keys())[i]

                            # 创建新的工作表
                            if sheet_name in temp_wb.sheetnames:
                                source_ws = temp_wb[sheet_name]
                                target_ws = final_wb.create_sheet(sheet_name)
                                copy_worksheet_content(source_ws, target_ws)

                            temp_wb.close()

                        # 保存最终文件
                        final_wb.save(out_file_path)
                        final_wb.close()
                    finally:
                        # 清理临时文件
                        for temp_path in temp_files:
                            if os.path.exists(temp_path):
                                os.remove(temp_path)
                processing_ring.update_status(ProgressStatus.SUCCESS, '拆分完成')
                if self.checkBox.value:
                    open_folder_in_explorer(output_folder_path_text.value)
            except Exception as e:
                processing_ring.update_status(ProgressStatus.ERROR, str(e))
            finally:
                header_tmp_folder_path = Path(output_folder_path_text.value, 'tmp')
                if header_tmp_folder_path.exists() and header_tmp_folder_path.is_dir():
                    shutil.rmtree(header_tmp_folder_path)

        # GUI
        if self.excel is None:
            return None
        else:
            processing_ring = ProgressRingComponent()
            split_config_component = ft.Column()
            for sheet_name in self.excel.sheets.keys():
                check_box = ft.Checkbox(label=sheet_name, value=False)
                header_rows = ft.TextField(label='请输入标题所占用的行数')
                split_column_index = ft.TextField(label='请输入根据第几列拆分')
                split_config_component.controls.append(ft.Row(controls=[check_box, header_rows,split_column_index],alignment=ft.MainAxisAlignment.CENTER,expand=True))

            button = ft.ElevatedButton(text='开始拆分', on_click=lambda _:_split_logic(output_folder_path_text,split_config_component,processing_ring))


            return ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Row(controls=[ft.Text("-------------配置拆分规则-------------")], alignment=ft.MainAxisAlignment.CENTER),
                        split_config_component,
                        ft.Row([button], alignment=ft.MainAxisAlignment.CENTER),
                        ft.Row([processing_ring], alignment=ft.MainAxisAlignment.CENTER),
                    ]
                ),
                margin=ft.Margin(left=0, top=20, right=0, bottom=0)
            )

    def _simple_split_excel(self, output_folder_path_text:ft.TextField):

        def redio_on_change(mode: ft.RadioGroup, button: ft.ElevatedButton, sheet_dropdown: ft.Dropdown,
                            columns_dropdown: ft.Dropdown):
            if mode.value == '0':
                sheet_dropdown.value = None
                columns_dropdown.value = None
                sheet_dropdown.options = []
                columns_dropdown.options = []
                sheet_dropdown.disabled = True
                columns_dropdown.disabled = True
                button.disabled = True
                sheet_dropdown.update()
                columns_dropdown.update()
                self.page.update()
            else:
                if self.excel:
                    sheet_dropdown.options = []
                    sheet_dropdown.options = [ft.DropdownOption(value) for value in self.excel.sheets.keys()]
                    sheet_dropdown.disabled = False
                    columns_dropdown.disabled = False
                    button.disabled = False
                    sheet_dropdown.update()
                    self.page.update()

        def sheet_dropdown_changed(selector: ft.Dropdown, columns_selector: ft.Dropdown):
            if columns_selector.value is not None:
                columns_selector.value = None
                columns_selector.visible = False
                sleep(0.01)
                columns_selector.visible = True
            columns_selector.options = [ft.DropdownOption(value) for value in self.excel.sheets[selector.value].columns]
            columns_selector.update()
            self.page.update()

        def business_logic(
                           mode: ft.RadioGroup,
                           sheet_selector: ft.Dropdown,
                           columns_selector: ft.Dropdown,
                           folder_path_text: ft.TextField,
                           progress: ProgressRingComponent
                           ):
            try:
                if self.excel is None:
                    raise RuntimeError('未提供待拆分文件或输出文件夹')
                progress.update_status(ProgressStatus.LOADING, '开始拆分')
                file_name = Path(self.excel.file_path).stem
                # 根据列拆分
                if mode.value == '1':
                    progress.update_status(ProgressStatus.LOADING, '开始读取源文件')
                    # df = pd.read_excel(self.excel.file_path, sheet_name=sheet_selector.value)
                    df = self.excel.sheets[sheet_selector.value].df_data
                    for group, data in df.groupby(columns_selector.value):
                        final_file_name = f'开始生成{file_name}_{group}.xlsx'
                        progress.update_status(ProgressStatus.LOADING, final_file_name)
                        data.to_excel(Path(folder_path_text.value, file_name + f"_{group}.xlsx"), index=False)
                elif mode.value == '0':
                    progress.update_status(ProgressStatus.LOADING, '开始读取源文件')
                    # excel = pd.ExcelFile(file_path_text.value)
                    for sheet_name in self.excel.sheets.keys():
                        final_file_name = f'开始生成{file_name}_{sheet_name}.xlsx'
                        progress.update_status(ProgressStatus.LOADING, final_file_name)
                        self.excel.sheets[sheet_name].df_data.to_excel(
                            Path(folder_path_text.value, file_name + f"_{sheet_name}.xlsx"), index=False)
                progress.update_status(ProgressStatus.SUCCESS, '完成文件拆分')
                open_folder_in_explorer(output_folder_path_text.value)
            except Exception as e:
                progress.update_status(ProgressStatus.ERROR, str(e))

        # 1.1列拆分功能
        sheet_selector = ft.Dropdown(
            label='请选择Sheet页',
            on_change=lambda _: sheet_dropdown_changed(sheet_selector, columns_selector),
            options=[],
            width=200,
            disabled=self.disable,
        )

        columns_selector = ft.Dropdown(
            label='请选择拆分列',
            options=[],
            width=300,
            disabled=self.disable,
        )

        # 拆分模式选择
        mode = ft.RadioGroup(
            ft.Row(
                [
                    ft.Radio(value="0", label="根据Sheet拆分"),
                    ft.Radio(value="1", label="根据列拆分"),
                ]
            ),
            value="0",
            on_change=lambda _: redio_on_change(mode, ft.Button(), sheet_selector, columns_selector)
        )
        # 处理进度条
        handel_progress = ProgressRingComponent()
        return ft.Column(
            controls=[
                ft.Row([ft.Text('选择拆分模式:'), mode], spacing=15),
                ft.Row([sheet_selector, columns_selector], alignment=ft.MainAxisAlignment.CENTER,
                       expand=True),
                ft.Row([ft.ElevatedButton(text='开始拆分',
                                          on_click=lambda _: business_logic(
                                              mode,
                                              sheet_selector,
                                              columns_selector,
                                              output_folder_path_text,
                                              handel_progress)
                                          )
                        ],
                       alignment=ft.MainAxisAlignment.CENTER,
                       expand=True
                       ),
                handel_progress
            ],  # 整个页面内容垂直居中
            spacing=30,
        )


    def gui(self):
        # 文件选择器组件
        file_path_text = ft.TextField(label='请选择需拆分文件(.xlsx)', expand=True)
        file_picker = ft.FilePicker(on_result=lambda e:self._picker(e,file_path_text))
        analyze_process_ring = ProgressRingComponent()
        analyze_button = ft.ElevatedButton(text='解析后拆分',
                                           on_click=lambda _: self._load_excel_file(file_path_text, analyze_process_ring, tab_page))
        # 输出文件夹选择器
        folder_picker = ft.FilePicker(on_result=lambda e: self._picker(e, folder_path_text, is_file=False))
        folder_path_text = ft.TextField(label='请选择输出文件夹', read_only=True, expand=True)
        excel_components = ft.Column(controls=[
            ft.Row([ft.IconButton(
                icon=ft.Icons.UPLOAD_FILE,
                on_click=lambda _: file_picker.pick_files()
            ),
                file_path_text,
            ]),
            ft.Row(
                [
                    ft.IconButton(
                        icon=ft.Icons.FOLDER,
                        on_click=lambda _: folder_picker.get_directory_path()),
                    folder_path_text
                ],
                expand=True
            ),
            self.checkBox,
            ft.Row(controls=[analyze_button,analyze_process_ring],alignment=ft.MainAxisAlignment.CENTER,expand=True)
        ])
        self.page.overlay.extend([file_picker, folder_picker])
        # 在二级Tab变更时更新二级Tab界面
        def on_excel_tab_change(_e, _output_folder_path_text:ft.TextField, _tabs:ft.Tabs):
            if _e.control.selected_index == 1:
                _tabs.tabs[1].content = self._split_multiple_headers_excel(_output_folder_path_text)
                _tabs.update()
                _e.page.update()
            pass

        excel_tab = ft.Tabs(
            on_change= lambda e: on_excel_tab_change(e, folder_path_text, excel_tab),
            selected_index=0,
            animation_duration=300,
            tabs=[
              ft.Tab(
                  text='无合并单元格单一表头拆分',
                  content=ft.Container(
                      content=ft.Container(),
                      padding=ft.padding.only(top=20),
                  ),
              ),
                ft.Tab(
                    text='复杂表头拆分',
                    content=ft.Container(
                        content=ft.Container(),
                        padding=ft.padding.only(top=20),
                    ),
                )
            ]
        )

        tab_page = ft.Tabs(
            visible= False,
            on_change= lambda e:self._tab_change(e,folder_path_text, excel_tab),
            selected_index=0,
            animation_duration=300,
            tabs=[
                ft.Tab(
                    text="基本拆分",
                    content=ft.Container(
                        content=self._simple_split_excel(folder_path_text),
                        padding=ft.padding.only(top=20),
                    ),
                ),
                ft.Tab(
                    text="高级拆分",
                    content=ft.Container(
                        content=excel_tab,
                        padding=ft.padding.only(top=20),  # 同样设置
                    )
                ),
            ]
        )
        return ft.Column(
            controls=[
                excel_components,
                tab_page
            ],
            spacing = 25,
        )




